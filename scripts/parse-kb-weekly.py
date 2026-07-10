# -*- coding: utf-8 -*-
"""KB부동산 주간시계열.xlsx 를 파싱해
1) 시/도 단위 최신 데이터 + 최근 12주 추이
2) 시/군/구 단위 최신 데이터 (시/도별로 그룹핑, geojson 코드에 매핑)
를 JSON으로 추출. 1회성 데이터 생성 스크립트 (커밋 대상 아님, 결과 JSON만 커밋).
"""
import glob
import json
import re
import openpyxl
from datetime import datetime


def find_latest(pattern):
    """kb_data/ 안에서 패턴에 맞는 파일 중 파일명이 가장 최근(날짜 접두어 기준)인 것을 선택.
    엑셀을 열어둘 때 생기는 ~$ 임시 잠금 파일은 제외."""
    matches = sorted(
        f for f in glob.glob(f"kb_data/{pattern}")
        if not f.split("\\")[-1].split("/")[-1].startswith("~$")
    )
    if not matches:
        raise FileNotFoundError(f"kb_data/{pattern} 에 해당하는 파일을 찾을 수 없습니다. KB부동산 데이터허브에서 새로 받은 파일을 kb_data/ 폴더에 넣어주세요.")
    return matches[-1]


SRC = find_latest("*주간시계열*.xlsx")
SRC_MONTHLY = find_latest("*월간*주택*시계열*.xlsx")
OUT_SIDO = "lib/data/kb-weekly-sido.json"
OUT_SGG = "lib/data/kb-weekly-sigungu.json"
OUT_INDEX_SIDO = "lib/data/kb-index-sido.json"
OUT_INDEX_SGG = "lib/data/kb-index-sigungu.json"
OUT_RATIO_SIDO = "lib/data/kb-ratio-sido.json"
OUT_RATIO_SGG = "lib/data/kb-ratio-sigungu.json"
GEOJSON_SIDO = "public/skorea-provinces.geojson"
GEOJSON_SGG = "public/skorea-municipalities.geojson"
WEEKS_10Y = 520
MONTHS_10Y = 120

CODE_TO_KB_NAME = {
    "11": "서울특별시", "21": "부산광역시", "22": "대구광역시", "23": "인천광역시",
    "24": "(구)광주광역시", "25": "대전광역시", "26": "울산광역시", "29": "세종특별자치시",
    "31": "경기도", "32": "강원특별자치도", "33": "충청북도", "34": "충청남도",
    "35": "전북특별자치도", "36": "(구)전라남도", "37": "경상북도", "38": "경상남도",
    "39": "제주특별자치도",
}
CODE_TO_KB_NAME_MONTHLY = {
    **CODE_TO_KB_NAME,
    "24": "광주광역시", "35": "전라북도", "36": "전라남도",
}
CODE_TO_DISPLAY_NAME = {
    "11": "서울", "21": "부산", "22": "대구", "23": "인천", "24": "광주",
    "25": "대전", "26": "울산", "29": "세종", "31": "경기", "32": "강원",
    "33": "충북", "34": "충남", "35": "전북", "36": "전남", "37": "경북",
    "38": "경남", "39": "제주",
}
# 집계용 그룹 헤더(실제 지역 아님) — 건너뜀
SKIP_HEADERS = {
    "강북14개구", "강남11개구", "6개광역시", "5개광역시", "수도권", "기타지방",
    "전남", "광주", "통합", "특별시",
}


def find_col(header, target):
    for i, name in enumerate(header):
        if name == target:
            return i
    for i, name in enumerate(header):
        if name and (name.startswith(target) or target.startswith(name)):
            return i
    raise KeyError(target)


def load_sheet_rows(ws):
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    header = rows[0]
    data_rows = [r for r in rows[2:] if r[0] is not None]
    return header, data_rows


DATE_STR_RE = re.compile(r"^'?\s*(\d{1,2})\s*\.\s*(\d{1,2})$")


def parse_monthly_dates(raw_col):
    """'98.12, '99.1 같은 명시적 표기 뒤에 2,3,4...로 이어지는 월간 시트 날짜 표기를 YYYY-MM으로 변환.
    데이터 영역 끝의 안내문 등 날짜가 아닌 행은 None으로 표시."""
    dates = []
    year = month = None
    for raw in raw_col:
        if raw is None:
            dates.append(None)
            continue
        if isinstance(raw, str):
            m = DATE_STR_RE.match(raw.strip())
            if not m:
                dates.append(None)
                continue
            yy, mm = int(m.group(1)), int(m.group(2))
            year = 1900 + yy if yy >= 50 else 2000 + yy
            month = mm
        elif isinstance(raw, (int, float)) and raw >= 100:
            # 연도 경계에서 'YYYY.1' 형태의 숫자로 표기되는 1월 (소수부는 신뢰하지 않음)
            year = int(raw)
            month = 1
        elif isinstance(raw, (int, float)) and year is not None:
            month = int(raw)  # 년도 변경 없는 이어지는 월 (1월은 항상 명시적으로 표기됨)
        else:
            dates.append(None)
            continue
        dates.append(f"{year:04d}-{month:02d}")
    return dates


def load_monthly_sheet_rows(ws):
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    header = rows[0]
    body = rows[2:]
    dates = parse_monthly_dates([r[0] for r in body])
    data_rows = [(date, *r[1:]) for date, r in zip(dates, body) if date is not None]
    return header, data_rows


def series_for(data_rows, idx, n=12, monthly=False):
    out = []
    for r in data_rows[-n:]:
        date, val = r[0], r[idx]
        if monthly:
            if not isinstance(val, (int, float)):
                continue
            out.append({"date": date, "value": round(val, 3)})
        else:
            if isinstance(date, datetime) and val is not None:
                out.append({"date": date.strftime("%Y-%m-%d"), "value": round(val, 3)})
    return out


def extract_sido(header, data_rows, n=12, monthly=False, name_map=None):
    name_map = name_map or CODE_TO_KB_NAME
    col_idx = {name: find_col(header, name) for name in [*name_map.values(), "전국"]}
    latest_row = data_rows[-1]
    latest_date = latest_row[0] if monthly else latest_row[0].strftime("%Y-%m-%d")

    nationwide_raw = latest_row[col_idx["전국"]]
    nationwide_latest = round(nationwide_raw, 3) if isinstance(nationwide_raw, (int, float)) else None
    nationwide_series = series_for(data_rows, col_idx["전국"], n, monthly)

    regions = []
    for code, kb_name in name_map.items():
        idx = col_idx[kb_name]
        latest_val = latest_row[idx]
        regions.append({
            "code": code,
            "name": CODE_TO_DISPLAY_NAME[code],
            "latest": round(latest_val, 3) if isinstance(latest_val, (int, float)) else None,
            "series": series_for(data_rows, idx, n, monthly),
        })

    return {
        "updatedAt": latest_date,
        "nationwide": {"latest": nationwide_latest, "series": nationwide_series},
        "regions": regions,
    }


def build_hierarchy(header, name_map=None):
    """헤더 목록을 순서대로 훑어서 (province_code, leaf_col_idx, matched_name) 리스트 생성."""
    name_map = name_map or CODE_TO_KB_NAME
    kb_name_to_code = {v: k for k, v in name_map.items()}
    leaves = []  # (province_code, matched_name_for_geojson, col_idx)
    current_province = None
    current_city = None  # (name, col_idx) — '시' 단위 부모
    city_has_child = set()  # col_idx of city columns that turned out to have gu children

    for i, name in enumerate(header):
        if not name or not isinstance(name, str):
            continue
        if name in ("구분", "전국"):
            continue
        if name in SKIP_HEADERS:
            continue

        if name in kb_name_to_code:
            current_province = kb_name_to_code[name]
            current_city = (name, i)  # 메트로 시/도 자신이 구의 부모 역할
            continue

        if current_province is None:
            continue

        if name.endswith("구") and not name[:-1].isdigit():
            parent_name = current_city[0] if current_city else name_map[current_province]
            matched = name if parent_name == name_map[current_province] else parent_name + name
            leaves.append((current_province, matched, i))
            if current_city:
                city_has_child.add(current_city[1])
            continue

        if name.endswith("군"):
            leaves.append((current_province, name, i))
            continue

        if name.endswith("시"):
            # 이전 '시'가 자식이 없었다면 그 자체로 leaf 등록
            if current_city and current_city[0] != name_map.get(current_province) \
                    and current_city[1] not in city_has_child:
                leaves.append((current_province, current_city[0], current_city[1]))
            current_city = (name, i)
            continue

    # 마지막 city 처리
    if current_city and current_city[0] != name_map.get(current_province) \
            and current_city[1] not in city_has_child:
        leaves.append((current_province, current_city[0], current_city[1]))

    return leaves


def extract_sigungu(header, data_rows, geo_names_by_prefix, n=12, monthly=False, name_map=None):
    leaves = build_hierarchy(header, name_map)
    latest_row = data_rows[-1]
    latest_date = latest_row[0] if monthly else latest_row[0].strftime("%Y-%m-%d")

    matched, unmatched = [], []
    for province_code, name, idx in leaves:
        candidates = geo_names_by_prefix.get(province_code, {})
        geo_code = candidates.get(name)
        val = latest_row[idx]
        if geo_code is None or not isinstance(val, (int, float)):
            unmatched.append(f"{province_code}:{name}")
            continue
        matched.append({
            "code": geo_code,
            "name": name,
            "latest": round(val, 3),
            "series": series_for(data_rows, idx, n, monthly),
        })

    return {"updatedAt": latest_date, "regions": matched}, unmatched


def main():
    print("주간 파일:", SRC)
    print("월간 파일:", SRC_MONTHLY)
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)

    # --- 시/도 ---
    h1, d1 = load_sheet_rows(wb["1.매매증감"])
    h2, d2 = load_sheet_rows(wb["2.전세증감"])
    sido_result = {"saleChange": extract_sido(h1, d1), "jeonseChange": extract_sido(h2, d2)}
    with open(OUT_SIDO, "w", encoding="utf-8") as f:
        json.dump(sido_result, f, ensure_ascii=False, indent=2)
    print("wrote", OUT_SIDO)

    # --- 시/군/구 ---
    geo = json.load(open(GEOJSON_SGG, encoding="utf-8"))
    geo_names_by_prefix = {}
    for feat in geo["features"]:
        code = feat["properties"]["code"]
        prefix = code[:2]
        geo_names_by_prefix.setdefault(prefix, {})[feat["properties"]["name"]] = code

    sale_sgg, sale_unmatched = extract_sigungu(h1, d1, geo_names_by_prefix)
    jeonse_sgg, jeonse_unmatched = extract_sigungu(h2, d2, geo_names_by_prefix)
    sgg_result = {"saleChange": sale_sgg, "jeonseChange": jeonse_sgg}
    with open(OUT_SGG, "w", encoding="utf-8") as f:
        json.dump(sgg_result, f, ensure_ascii=False, indent=2)
    print("wrote", OUT_SGG)
    print("sale matched:", len(sale_sgg["regions"]), "unmatched:", sorted(set(sale_unmatched)))
    print("jeonse matched:", len(jeonse_sgg["regions"]), "unmatched:", sorted(set(jeonse_unmatched)))

    # --- 매매지수/전세지수 (최근 10년 주간) ---
    h3, d3 = load_sheet_rows(wb["3.매매지수"])
    h4, d4 = load_sheet_rows(wb["4.전세지수"])

    index_sido_result = {
        "saleIndex": extract_sido(h3, d3, WEEKS_10Y),
        "jeonseIndex": extract_sido(h4, d4, WEEKS_10Y),
    }
    with open(OUT_INDEX_SIDO, "w", encoding="utf-8") as f:
        json.dump(index_sido_result, f, ensure_ascii=False, indent=2)
    print("wrote", OUT_INDEX_SIDO)

    sale_idx_sgg, _ = extract_sigungu(h3, d3, geo_names_by_prefix, WEEKS_10Y)
    jeonse_idx_sgg, _ = extract_sigungu(h4, d4, geo_names_by_prefix, WEEKS_10Y)
    index_sgg_result = {"saleIndex": sale_idx_sgg, "jeonseIndex": jeonse_idx_sgg}
    with open(OUT_INDEX_SGG, "w", encoding="utf-8") as f:
        json.dump(index_sgg_result, f, ensure_ascii=False, indent=2)
    print("wrote", OUT_INDEX_SGG)

    # --- 아파트 매매전세비 (월간, 최근 10년) ---
    wb_m = openpyxl.load_workbook(SRC_MONTHLY, read_only=True, data_only=True)
    h5, d5 = load_monthly_sheet_rows(wb_m["28.아파트매매전세비"])

    ratio_sido_result = {"ratio": extract_sido(h5, d5, MONTHS_10Y, monthly=True, name_map=CODE_TO_KB_NAME_MONTHLY)}
    with open(OUT_RATIO_SIDO, "w", encoding="utf-8") as f:
        json.dump(ratio_sido_result, f, ensure_ascii=False, indent=2)
    print("wrote", OUT_RATIO_SIDO)

    geo_m = json.load(open(GEOJSON_SGG, encoding="utf-8"))
    geo_names_by_prefix_m = {}
    for feat in geo_m["features"]:
        code = feat["properties"]["code"]
        geo_names_by_prefix_m.setdefault(code[:2], {})[feat["properties"]["name"]] = code

    ratio_sgg, ratio_unmatched = extract_sigungu(h5, d5, geo_names_by_prefix_m, MONTHS_10Y, monthly=True, name_map=CODE_TO_KB_NAME_MONTHLY)
    ratio_sgg_result = {"ratio": ratio_sgg}
    with open(OUT_RATIO_SGG, "w", encoding="utf-8") as f:
        json.dump(ratio_sgg_result, f, ensure_ascii=False, indent=2)
    print("wrote", OUT_RATIO_SGG)
    print("ratio matched:", len(ratio_sgg["regions"]), "unmatched:", sorted(set(ratio_unmatched)))


if __name__ == "__main__":
    main()
